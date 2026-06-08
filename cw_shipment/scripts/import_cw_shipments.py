#!/usr/bin/env python3
"""
Import shipments from CW Shipment Log.xlsx into an Odoo 19 database via XML-RPC.

Usage:
    python import_cw_shipments.py \
        --url http://localhost:8069 \
        --db CW19_Test \
        --user admin --password ADMIN \
        --file "CW Shipment Log.xlsx" \
        [--sheet 工作表1] [--dry-run] [--update] [--limit N]

Behavior:
    * Reads the XLSX, normalizes values (case, container sizes, port spellings).
    * Groups rows by PI No. — rows sharing a PI become container lines of one shipment.
    * Looks up partners (customer, supplier, forwarder, carrier) in the target DB by
      normalized name; only creates them if missing.
    * Forwarders and carriers are tagged with the "Service Provider" partner category
      (and sub-tags "Forwarder" / "Carrier" if the cw_shipment module is installed).
    * --dry-run prints what would happen, makes no writes.
    * --update overwrites existing shipments (matched by PI No.); default skips them.
"""

from __future__ import annotations

import argparse
import collections
import datetime as dt
import logging
import re
import sys
import xmlrpc.client
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl is required. Install with: pip install openpyxl", file=sys.stderr)
    sys.exit(2)


log = logging.getLogger("cw_shipment.import")


# ---------------------------------------------------------------------------
# Header schema (row 3 of the spreadsheet, 1-indexed column numbers)
# ---------------------------------------------------------------------------
COL = {
    'pi_no':              1,   # A
    'customer':           2,   # B
    'term':               3,   # C
    'supplier':           4,   # D
    'forwarder':          5,   # E
    'freight_cost':       6,   # F
    'cargo_ready_date':   7,   # G
    'inspection_date':    8,   # H
    'status':             9,   # I
    'doc_to_customer':   10,   # J
    'telex_released':    11,   # K
    'port_of_loading':   12,   # L
    'description':       13,   # M
    'ctn_pkg':           14,   # N
    'gross_weight':      15,   # O
    'total_cbm':         16,   # P
    'cntr_size':         17,   # Q
    'vessel_name':       18,   # R
    'open_date':         19,   # S
    'closing_date':      20,   # T
    'si_cutoff':         21,   # U
    'vgm_cutoff':        22,   # V
    'etd':               23,   # W
    'eta_original':      24,   # X
    'eta':               25,   # Y
    'destination':       26,   # Z
    'carrier':           27,   # AA
    'cntr_number':       28,   # AB
    'bl_number':         29,   # AC
    'invoice_no':        30,   # AD
    'invoice_amount':    31,   # AE
    'pi_no_alt':         32,   # AF
    'fty_pi':            33,   # AG
    'sent_doc':          34,   # AH
    'courier':           35,   # AI
    'remarks':           36,   # AJ
}

HEADER_ROW = 3
FIRST_DATA_ROW = 4


# ---------------------------------------------------------------------------
# Normalization
# ---------------------------------------------------------------------------
def norm_name(value):
    """Normalize a partner / port name: trim, collapse whitespace, title-case unless already mixed."""
    if value is None:
        return None
    s = str(value).strip()
    if not s:
        return None
    s = re.sub(r'\s+', ' ', s)
    # If the string is ALL CAPS or all lowercase, title-case it. Otherwise preserve mixed case.
    if s.isupper() or s.islower():
        s = s.title()
    return s


# Manual aliases for known-equivalent names. Add entries here when reviewing
# the dry-run report flags new duplicates.
CUSTOMER_ALIASES = {
    'Bashco': 'Bashco',
    'BASHCO': 'Bashco',
    'JADA': 'JADA',
    'ICC': 'ICC',
    'BJS': 'BJS',
    'BJ Inv.': 'BJS',  # likely the same per spreadsheet pattern
    'Wys/Wyg Ent.': 'WYS/WYG Ent.',
    'Eventify': 'Eventify',
    'Mane': 'Mane',
    'Chickmont': 'Chickmont',
    'Super Feeds': 'Super Feeds',
    'Lucky Store': 'Lucky Store',
    'Ridgelands': 'Ridgelands',
    'Western Wholesales': 'Western Wholesales',
    'Margaritaville': 'Margaritaville',
    'Simplex': 'SIMPLEX',
    'Prudential Belife': 'Prudential Belife',
    'Rainforest': 'Rainforest',
}

PORT_ALIASES = {
    'Ningbo': 'Ningbo',
    'NINGBO': 'Ningbo',
    'Qingdao': 'Qingdao',
    'QINGDAO': 'Qingdao',
    'Shanghai': 'Shanghai',
    'Shenzhen': 'Shenzhen',
    'Yantian': 'Yantian',
    'Shekou': 'Shekou',
    'Nansha GZ': 'Nansha',
    'NANSHA GZ': 'Nansha',
    'Guangzhou': 'Guangzhou',
    'GUANGZHOU': 'Guangzhou',
    'Xingang, Tianjin': 'Xingang, Tianjin',
    'Tainjin': 'Xingang, Tianjin',  # typo in source
    'Nanjing': 'Nanjing',
    'Nanjin / Shanghai': 'Nanjing',
    'Xiamen': 'Xiamen',
    'Hong Kong': 'Hong Kong',
    'HONG KONG': 'Hong Kong',
    'Ningbo Yiwu': 'Ningbo Yiwu',
}

DESTINATION_ALIASES = {
    'Kingston': 'Kingston, Jamaica',
    'KINGSTON, JAMAICA': 'Kingston, Jamaica',
    'Kingston, Jamaica': 'Kingston, Jamaica',
    'Barbados': 'Bridgetown, Barbados',
    'BRIDGETOWN, BARBADOS': 'Bridgetown, Barbados',
    'Bridgetown': 'Bridgetown, Barbados',
    'BRIDGETOWN': 'Bridgetown, Barbados',
    'Bridgeown': 'Bridgetown, Barbados',  # typo
    'Rijeka': 'Rijeka, Croatia',
    'RIJEKA, CROATIA': 'Rijeka, Croatia',
    'JEBEL ALI': 'Jebel Ali, UAE',
    'Suriname': 'Paramaribo, Suriname',
    'Paramaribo, Suriname': 'Paramaribo, Suriname',
    'Miami': 'Miami, USA',
    'DXB': 'Jebel Ali, UAE',
    'ABJ': 'Abidjan, Cote d\'Ivoire',
    'NINGBO': 'Ningbo',  # appears once in dest column — likely data entry error, leave as-is port name
}

STATUS_MAP = {
    'shipped': 'shipped',
    'SHIPPED': 'shipped',
}

# Carrier names are well-known shipping-line acronyms; keep them uppercase
# (or in the company's own canonical casing) rather than title-casing.
CARRIER_ALIASES = {
    'APL': 'APL',
    'Apl': 'APL',
    'CMA': 'CMA CGM',
    'Cma': 'CMA CGM',
    'CMA CGM': 'CMA CGM',
    'Cma Cgm': 'CMA CGM',
    'COSCO': 'COSCO',
    'Cosco': 'COSCO',
    'EMC': 'EMC',
    'Emc': 'EMC',
    'EUMEX': 'EUMEX',
    'Eumex': 'EUMEX',
    'HPL': 'Hapag-Lloyd',
    'Hpl': 'Hapag-Lloyd',
    'MAERSK': 'Maersk',
    'Maersk': 'Maersk',
    'MSK': 'Maersk',
    'Msk': 'Maersk',
    'MSC': 'MSC',
    'Msc': 'MSC',
    'ZIM': 'ZIM',
    'Zim': 'ZIM',
}

# Map raw container-size strings to (selection_value, count_of_containers)
CONTAINER_SIZE_MAP = {
    "1X20'": ('20', 1),
    "1x20'": ('20', 1),
    "1X20' SOC": ('20_soc', 1),
    "1X40'": ('40', 1),
    "1x40'": ('40', 1),
    "1X40'HQ": ('40hq', 1),
    "1x40'HQ": ('40hq', 1),
    "1X40HQ": ('40hq', 1),
    "1x40HQ": ('40hq', 1),
    "1x40HQ'": ('40hq', 1),
    "2X40'HQ": ('40hq', 2),
    "2x40'HQ": ('40hq', 2),
    "3X40'HQ": ('40hq', 3),
    "4X40'HQ": ('40hq', 4),
    "4x40'HQ": ('40hq', 4),
    "PART OF 40'HQ": ('partial_40hq', 1),
    "LCL": ('lcl', 1),
    "LCL/CY": ('lcl_cy', 1),
    "AIR": ('air', 1),
}


def parse_container_size(raw):
    """Return (selection_value, n_containers, raw_kept)."""
    if not raw:
        return ('40hq', 1, None)
    key = re.sub(r'\s+', '', str(raw).strip()).upper()
    # Re-add apostrophes? Easier: index by uppercased+stripped, with table normalized the same way.
    for k, v in CONTAINER_SIZE_MAP.items():
        if re.sub(r'\s+', '', k).upper() == key:
            return (v[0], v[1], str(raw).strip())
    log.warning("Unmapped container size: %r — falling back to 'other'", raw)
    return ('other', 1, str(raw).strip())


def to_date(value):
    """Convert openpyxl cell value to ISO date string (or None)."""
    if value is None or value == '':
        return None
    if isinstance(value, dt.datetime):
        return value.date().isoformat()
    if isinstance(value, dt.date):
        return value.isoformat()
    # Sometimes Excel stores dates as strings.
    s = str(value).strip()
    for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y', '%d-%b-%y', '%d-%b-%Y'):
        try:
            return dt.datetime.strptime(s, fmt).date().isoformat()
        except ValueError:
            continue
    log.debug("Could not parse date: %r", value)
    return None


def to_float(value):
    if value is None or value == '':
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    # Keep digits, decimal point, leading minus only. Drop trailing punctuation like ".-".
    s = str(value).strip()
    m = re.match(r'-?\d+(?:\.\d+)?', s.replace(',', ''))
    return float(m.group(0)) if m else 0.0


def to_int(value):
    if value is None or value == '':
        return 0
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return int(value)
    m = re.match(r'-?\d+', str(value).strip().replace(',', ''))
    return int(m.group(0)) if m else 0


def to_str(value):
    if value is None:
        return None
    s = str(value).strip()
    return s or None


# ---------------------------------------------------------------------------
# Parse XLSX into row dicts
# ---------------------------------------------------------------------------
def parse_xlsx(path, sheet_name=None):
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb[sheet_name] if sheet_name else wb[wb.sheetnames[0]]

    rows = []
    for row_idx in range(FIRST_DATA_ROW, ws.max_row + 1):
        pi = ws.cell(row=row_idx, column=COL['pi_no']).value
        if pi is None or str(pi).strip() == '':
            continue

        row = {key: ws.cell(row=row_idx, column=col).value for key, col in COL.items()}
        row['_row_number'] = row_idx
        rows.append(row)

    return rows


# ---------------------------------------------------------------------------
# Group rows by PI into (header_dict, [container_dicts]) tuples
# ---------------------------------------------------------------------------
def build_shipments(rows):
    """Group rows by PI No. into shipments with container lines."""
    by_pi = collections.OrderedDict()

    for row in rows:
        pi = to_str(row['pi_no'])
        if not pi:
            continue

        bucket = by_pi.setdefault(pi, {'header': None, 'lines': []})

        # Header is taken from the first occurrence; subsequent rows only contribute container lines.
        if bucket['header'] is None:
            bucket['header'] = build_header(row, pi)

        for line in build_lines(row):
            bucket['lines'].append(line)

    return by_pi


def build_header(row, pi):
    customer_raw = to_str(row['customer'])
    supplier_raw = to_str(row['supplier'])
    forwarder_raw = to_str(row['forwarder'])
    port_raw = to_str(row['port_of_loading'])
    dest_raw = to_str(row['destination'])
    status_raw = to_str(row['status'])

    return {
        'pi_no': pi,
        'fty_pi': to_str(row['fty_pi']),
        'customer_name': resolve_alias(CUSTOMER_ALIASES, customer_raw),
        'supplier_name': norm_name(supplier_raw),
        'forwarder_name': norm_name(forwarder_raw),
        'term': (to_str(row['term']) or 'FOB').upper(),
        'freight_cost': to_float(row['freight_cost']),
        'invoice_amount': to_float(row['invoice_amount']),
        'invoice_no': to_str(row['invoice_no']),
        'cargo_ready_date': to_date(row['cargo_ready_date']),
        'inspection_date': to_date(row['inspection_date']),
        'doc_to_customer_date': to_date(row['doc_to_customer']),
        'telex_released_date': to_date(row['telex_released']),
        'sent_doc_date': to_date(row['sent_doc']),
        'courier': to_str(row['courier']),
        'description': to_str(row['description']),
        'ctn_pkg': to_int(row['ctn_pkg']),
        'gross_weight': to_float(row['gross_weight']),
        'total_cbm': to_float(row['total_cbm']),
        'port_of_loading_name': resolve_alias(PORT_ALIASES, port_raw),
        'destination_port_name': resolve_alias(DESTINATION_ALIASES, dest_raw),
        'status': STATUS_MAP.get(status_raw, 'draft') if status_raw else 'draft',
        'remarks': to_str(row['remarks']),
        '_raw': {  # preserved for import_raw field
            'row_number': row['_row_number'],
            'customer': customer_raw,
            'supplier': supplier_raw,
            'forwarder': forwarder_raw,
            'port': port_raw,
            'destination': dest_raw,
            'status': status_raw,
            'term': to_str(row['term']),
            'cntr_size': to_str(row['cntr_size']),
        },
    }


def build_lines(row):
    """Build container lines from a single spreadsheet row.

    If size says "2x40HQ", we create 2 lines (CNTR No. only attaches to the first
    unless comma-separated). Comma-separated CNTR No.s are split into multiple lines.
    """
    size_raw = to_str(row['cntr_size'])
    size_value, n_containers, size_kept = parse_container_size(size_raw)

    cntr_raw = to_str(row['cntr_number'])
    cntr_numbers = []
    if cntr_raw:
        cntr_numbers = [c.strip() for c in re.split(r'[,;/]', cntr_raw) if c.strip()]

    # If multiple CNTR Nos. provided, use that count; otherwise rely on parsed n.
    actual_n = max(n_containers, len(cntr_numbers)) if cntr_numbers else n_containers

    lines = []
    for i in range(actual_n):
        cntr_number = cntr_numbers[i] if i < len(cntr_numbers) else None
        carrier_raw = to_str(row['carrier'])
        lines.append({
            'container_size': size_value,
            'container_size_raw': size_kept,
            'cntr_number': cntr_number,
            'bl_number': to_str(row['bl_number']),
            'carrier_name': resolve_alias(CARRIER_ALIASES, carrier_raw),
            'vessel_name': to_str(row['vessel_name']),
            'open_date': to_date(row['open_date']),
            'closing_date': to_date(row['closing_date']),
            'si_cutoff': to_date(row['si_cutoff']),
            'vgm_cutoff': to_date(row['vgm_cutoff']),
            'etd': to_date(row['etd']),
            'eta_original': to_date(row['eta_original']),
            'eta': to_date(row['eta']),
            'sequence': 10 + (i * 10),
        })
    return lines


def resolve_alias(alias_map, value):
    if value is None:
        return None
    if value in alias_map:
        return alias_map[value]
    normed = norm_name(value)
    if normed in alias_map:
        return alias_map[normed]
    return normed


# ---------------------------------------------------------------------------
# Odoo XML-RPC client
# ---------------------------------------------------------------------------
class OdooClient:
    def __init__(self, url, db, user, password, dry_run=False):
        self.url = url.rstrip('/')
        self.db = db
        self.user = user
        self.password = password
        self.dry_run = dry_run
        self.uid = None
        self._models = None
        self._partner_by_name = {}
        self._port_by_name = {}
        self._cat_by_name = {}
        self._existing_pi = set()

    def connect(self):
        common = xmlrpc.client.ServerProxy(f"{self.url}/xmlrpc/2/common", allow_none=True)
        version = common.version()
        log.info("Connected to %s — Odoo %s", self.url, version.get('server_version'))
        self.uid = common.authenticate(self.db, self.user, self.password, {})
        if not self.uid:
            raise RuntimeError(f"Authentication failed for {self.user}@{self.db}")
        log.info("Authenticated as uid=%s on db=%s", self.uid, self.db)
        self._models = xmlrpc.client.ServerProxy(f"{self.url}/xmlrpc/2/object", allow_none=True)

    def execute(self, model, method, args, kwargs=None):
        return self._models.execute_kw(self.db, self.uid, self.password, model, method, args, kwargs or {})

    # ----- pre-loading caches ----------------------------------------------
    def preload(self):
        log.info("Preloading partner names…")
        partners = self.execute('res.partner', 'search_read',
                                [[]],
                                {'fields': ['id', 'name', 'customer_rank', 'supplier_rank', 'category_id']})
        for p in partners:
            key = (p['name'] or '').strip().lower()
            if key:
                self._partner_by_name.setdefault(key, []).append(p)
        log.info("  %d partners loaded", len(partners))

        log.info("Preloading ports…")
        try:
            ports = self.execute('cw.port', 'search_read', [[]], {'fields': ['id', 'name']})
            for p in ports:
                self._port_by_name[p['name'].strip().lower()] = p['id']
            log.info("  %d ports loaded", len(ports))
        except xmlrpc.client.Fault as exc:
            log.warning("cw.port model not available — module may not be installed: %s", exc.faultString)

        log.info("Preloading partner categories…")
        cats = self.execute('res.partner.category', 'search_read',
                            [[]], {'fields': ['id', 'name']})
        for c in cats:
            self._cat_by_name[c['name'].strip().lower()] = c['id']

        log.info("Preloading existing shipments…")
        try:
            existing = self.execute('cw.shipment', 'search_read', [[]], {'fields': ['pi_no']})
            self._existing_pi = {s['pi_no'] for s in existing}
            log.info("  %d shipments already in DB", len(self._existing_pi))
        except xmlrpc.client.Fault as exc:
            log.warning("cw.shipment model not available — module may not be installed: %s", exc.faultString)

    # ----- partner lookup / create -----------------------------------------
    def get_or_create_partner(self, name, role):
        """role: 'customer' | 'supplier' | 'forwarder' | 'carrier'"""
        if not name:
            return None

        key = name.strip().lower()
        matches = self._partner_by_name.get(key, [])

        if matches:
            return matches[0]['id']

        # Build vals according to role
        vals = {'name': name, 'is_company': True}
        category_ids = []

        if role == 'customer':
            vals['customer_rank'] = 1
        elif role == 'supplier':
            vals['supplier_rank'] = 1
        elif role == 'forwarder':
            cat_sp = self._cat_by_name.get('service provider')
            cat_fw = self._cat_by_name.get('forwarder')
            if cat_sp:
                category_ids.append(cat_sp)
            if cat_fw:
                category_ids.append(cat_fw)
        elif role == 'carrier':
            cat_sp = self._cat_by_name.get('service provider')
            cat_c = self._cat_by_name.get('carrier')
            if cat_sp:
                category_ids.append(cat_sp)
            if cat_c:
                category_ids.append(cat_c)
        if category_ids:
            vals['category_id'] = [(6, 0, category_ids)]

        if self.dry_run:
            log.info("    [DRY] would create partner %r (role=%s)", name, role)
            return None

        new_id = self.execute('res.partner', 'create', [vals])
        log.info("    created partner %r (role=%s) id=%s", name, role, new_id)
        # Update cache
        self._partner_by_name.setdefault(key, []).append({
            'id': new_id, 'name': name,
            'customer_rank': vals.get('customer_rank', 0),
            'supplier_rank': vals.get('supplier_rank', 0),
            'category_id': category_ids,
        })
        return new_id

    def get_or_create_port(self, name):
        if not name:
            return None
        key = name.strip().lower()
        if key in self._port_by_name:
            return self._port_by_name[key]
        if self.dry_run:
            log.info("    [DRY] would create port %r", name)
            return None
        new_id = self.execute('cw.port', 'create', [{'name': name}])
        log.info("    created port %r id=%s", name, new_id)
        self._port_by_name[key] = new_id
        return new_id

    # ----- shipment write --------------------------------------------------
    def upsert_shipment(self, header, lines, partner_ids, port_ids, update_existing=False):
        pi = header['pi_no']

        line_vals_list = []
        for line in lines:
            carrier_id = partner_ids.get(('carrier', line['carrier_name'])) if line['carrier_name'] else False
            line_vals_list.append((0, 0, {
                'container_size': line['container_size'],
                'container_size_raw': line['container_size_raw'],
                'cntr_number': line['cntr_number'],
                'bl_number': line['bl_number'],
                'carrier_id': carrier_id,
                'vessel_name': line['vessel_name'],
                'open_date': line['open_date'],
                'closing_date': line['closing_date'],
                'si_cutoff': line['si_cutoff'],
                'vgm_cutoff': line['vgm_cutoff'],
                'etd': line['etd'],
                'eta_original': line['eta_original'],
                'eta': line['eta'],
                'sequence': line['sequence'],
            }))

        vals = {
            'pi_no': pi,
            'fty_pi': header['fty_pi'],
            'customer_id': partner_ids.get(('customer', header['customer_name'])) or False,
            'supplier_id': partner_ids.get(('supplier', header['supplier_name'])) or False,
            'forwarder_id': partner_ids.get(('forwarder', header['forwarder_name'])) or False,
            'term': header['term'] if header['term'] in ('FOB', 'CIF', 'EXW', 'FCA', 'CFR', 'DAP', 'DDP') else 'FOB',
            'freight_cost': header['freight_cost'],
            'invoice_amount': header['invoice_amount'],
            'invoice_no': header['invoice_no'],
            'cargo_ready_date': header['cargo_ready_date'],
            'inspection_date': header['inspection_date'],
            'doc_to_customer_date': header['doc_to_customer_date'],
            'telex_released_date': header['telex_released_date'],
            'sent_doc_date': header['sent_doc_date'],
            'courier': header['courier'],
            'description': header['description'],
            'ctn_pkg': header['ctn_pkg'],
            'gross_weight': header['gross_weight'],
            'total_cbm': header['total_cbm'],
            'port_of_loading_id': port_ids.get(header['port_of_loading_name']) or False,
            'destination_port_id': port_ids.get(header['destination_port_name']) or False,
            'status': header['status'],
            'remarks': header['remarks'],
            'import_raw': repr(header['_raw']),
            'container_ids': line_vals_list,
        }

        if pi in self._existing_pi:
            if not update_existing:
                log.info("  SKIP existing shipment PI=%s", pi)
                return None
            # Update path: find id, write header, replace lines
            existing = self.execute('cw.shipment', 'search', [[['pi_no', '=', pi]]])
            if existing:
                ship_id = existing[0]
                if self.dry_run:
                    log.info("  [DRY] would UPDATE shipment PI=%s id=%s (%d lines)", pi, ship_id, len(line_vals_list))
                    return ship_id
                # Clear existing lines first
                line_ids = self.execute('cw.shipment.container', 'search', [[['shipment_id', '=', ship_id]]])
                if line_ids:
                    self.execute('cw.shipment.container', 'unlink', [line_ids])
                self.execute('cw.shipment', 'write', [[ship_id], vals])
                log.info("  UPDATED shipment PI=%s id=%s (%d containers)", pi, ship_id, len(line_vals_list))
                return ship_id

        # Create path
        if self.dry_run:
            log.info("  [DRY] would CREATE shipment PI=%s (%d containers)", pi, len(line_vals_list))
            return None
        new_id = self.execute('cw.shipment', 'create', [vals])
        log.info("  CREATED shipment PI=%s id=%s (%d containers)", pi, new_id, len(line_vals_list))
        self._existing_pi.add(pi)
        return new_id


# ---------------------------------------------------------------------------
# Main flow
# ---------------------------------------------------------------------------
def collect_partners_and_ports(shipments):
    """Return (set of (role, name), set of port_names)."""
    partners = set()
    ports = set()
    for pi, bucket in shipments.items():
        h = bucket['header']
        if h['customer_name']:
            partners.add(('customer', h['customer_name']))
        if h['supplier_name']:
            partners.add(('supplier', h['supplier_name']))
        if h['forwarder_name']:
            partners.add(('forwarder', h['forwarder_name']))
        if h['port_of_loading_name']:
            ports.add(h['port_of_loading_name'])
        if h['destination_port_name']:
            ports.add(h['destination_port_name'])
        for line in bucket['lines']:
            if line['carrier_name']:
                partners.add(('carrier', line['carrier_name']))
    return partners, ports


def run(args):
    logging.basicConfig(
        level=logging.DEBUG if args.verbose else logging.INFO,
        format='%(asctime)s %(levelname)-7s %(message)s',
        datefmt='%H:%M:%S',
    )

    path = Path(args.file).expanduser().resolve()
    if not path.exists():
        log.error("File not found: %s", path)
        return 2

    log.info("=== STAGE 1: parse XLSX ===")
    rows = parse_xlsx(path, args.sheet)
    log.info("  %d non-empty data rows", len(rows))

    shipments = build_shipments(rows)
    if args.limit:
        shipments = collections.OrderedDict(list(shipments.items())[:args.limit])
    log.info("  %d distinct shipments (by PI No.)", len(shipments))

    partner_keys, port_names = collect_partners_and_ports(shipments)
    log.info("  %d distinct partners to resolve", len(partner_keys))
    log.info("  %d distinct ports to resolve", len(port_names))

    if args.parse_only:
        log.info("=== --parse-only — done ===")
        return 0

    log.info("=== STAGE 2: connect to Odoo ===")
    client = OdooClient(args.url, args.db, args.user, args.password, dry_run=args.dry_run)
    client.connect()
    client.preload()

    log.info("=== STAGE 3: resolve partners ===")
    partner_ids = {}
    new_partners = 0
    for role, name in sorted(partner_keys):
        pid = client.get_or_create_partner(name, role)
        if pid:
            partner_ids[(role, name)] = pid
        elif client.dry_run:
            partner_ids[(role, name)] = None
            new_partners += 1

    log.info("=== STAGE 4: resolve ports ===")
    port_ids = {}
    for name in sorted(port_names):
        port_ids[name] = client.get_or_create_port(name)

    log.info("=== STAGE 5: create shipments ===")
    created, updated, skipped = 0, 0, 0
    for pi, bucket in shipments.items():
        header = bucket['header']

        # Note shipments with missing customer or supplier — import them anyway
        # since the source data has some legacy PIs without these.
        gaps = []
        if not header['customer_name']:
            gaps.append('customer')
        if not header['supplier_name']:
            gaps.append('supplier')
        if gaps:
            log.warning("  PI=%s — importing with missing %s (fill in via Odoo later)", pi, ', '.join(gaps))

        try:
            result = client.upsert_shipment(header, bucket['lines'], partner_ids, port_ids, update_existing=args.update)
            if result is None and pi in client._existing_pi and not args.update:
                skipped += 1
            elif pi in client._existing_pi and args.update:
                updated += 1
            else:
                created += 1
        except xmlrpc.client.Fault as exc:
            log.error("  PI=%s FAILED: %s", pi, exc.faultString)
            skipped += 1

    log.info("=== SUMMARY ===")
    log.info("  Shipments:  %d created, %d updated, %d skipped", created, updated, skipped)
    if args.dry_run:
        log.info("  (dry-run — nothing was written to %s)", args.db)
    return 0


def main():
    p = argparse.ArgumentParser(description="Import CW Shipment Log into Odoo 19")
    p.add_argument('--url', default='http://localhost:8069', help='Odoo URL')
    p.add_argument('--db', required=True, help='Database name (e.g. CW19_Test)')
    p.add_argument('--user', default='admin', help='Login (default: admin)')
    p.add_argument('--password', required=True, help='Password or API key')
    p.add_argument('--file', required=True, help='Path to CW Shipment Log.xlsx')
    p.add_argument('--sheet', default=None, help='Sheet name (default: first sheet)')
    p.add_argument('--dry-run', action='store_true', help='No writes; report what would happen')
    p.add_argument('--update', action='store_true', help='Update existing shipments matched by PI No.')
    p.add_argument('--limit', type=int, default=0, help='Import at most N shipments (for testing)')
    p.add_argument('--parse-only', action='store_true', help='Just parse the file, do not connect to Odoo')
    p.add_argument('-v', '--verbose', action='store_true')
    args = p.parse_args()
    return run(args)


if __name__ == '__main__':
    sys.exit(main())
